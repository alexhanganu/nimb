'''
calculating the T-test for the means of two independent samples of scores (ttest_ind)
kurtosis for the column with Fisher definition, normal ==>0
skew: for normally distributed data, the skewness should be about zero. If >0, means more weight in the right tail of the distribution.
'''

import os
import string
import itertools

import pandas as pd
import numpy as np
from scipy import stats

from stats.db_processing import Table
from distribution import utilities


def combinations_get(ls, lvl=0):
    """combining values from ls
    Args:
        ls = initial list() with values to be combined (ex : val du gr1 et val du gr 4)
        lvl = int() of number of values to be combined, (
                0 = blank tuple,
                1 = tuple with 1 value,
                2 = tuples of 2 values
    Return:
        combined = final list() that containes tuples() with combinations
                    exemple: [(group1, group2), (group2, group3), (group 1, group3)]
    """
    if lvl == 0:
        return [tuple()]
    elif lvl == 1:
        return [(i,) for i in ls]
    elif lvl == 2:
        result = list()
        combined = list(itertools.product(ls, ls))
        combined_diff = [i for i in combined if i[0] != i[1]]
        for i in combined_diff:
            if i not in result and (i[1], i[0]) not in result:
                result.append(i)
        return result
    else:
        print("    requests for combinations higher than 2\
            cannot be performed because FreeSurfer does not take them")



class ttest_do():
    def __init__(self, df, group_col, ls_cols, groups, path_save_res, p_thresh = 0.05):
        self.df = df
        self.group_col = group_col
        self.ls_cols = ls_cols
        self.groups = groups
        self.path_save_res = path_save_res
        self.ls_meas = get_names_of_measurements()
        self.ls_struct = get_names_of_structures()
        self.res_ttest = self.compute_ttest_for_col(p_thresh)
        self.tab       = Table()

    def compute_ttest_for_col(self, p_thresh):
        res_4df = {'features':[], 'ttest':[], 'welch':[]}
        res = dict()
        for col in self.ls_cols:
            group1 = self.df[self.df[self.group_col] == self.groups[0]][col]
            group2 = self.df[self.df[self.group_col] == self.groups[1]][col]
            ttest_eq_pop_var = stats.ttest_ind(group1, group2, equal_var=True)
            ttest_welch = stats.ttest_ind(group1, group2, equal_var=False)
            if ttest_eq_pop_var[1] < p_thresh:
                meas, struct = get_structure_measurement(col, self.ls_meas, self.ls_struct)
                #print('{:<15} {}'.format(meas, struct))
                res[col] = {'{}, mean'.format(self.groups[0]): stats.tmean(group1),
                            '{}, std'.format(self.groups[1]): stats.tstd(group2),
                            '{}, mean'.format(self.groups[1]): stats.tmean(group2),
                            '{}, std'.format(self.groups[1]): stats.tstd(group2),
                            'ttest':ttest_eq_pop_var[1],
                            'welch':ttest_welch[1],
                            'kurtosis':stats.kurtosis(self.df[self.group_col]),
                            'skewness':stats.skew(self.df[self.group_col])}
                res_4df['features'].append(struct+' ('+meas+')')
                res_4df['ttest'].append(ttest_eq_pop_var[1])
                res_4df['welch'].append(ttest_welch[1])
        self.save_res(res_4df)
        return res

    def save_res(self, res_4df):
        df_result = self.tab.create_df_from_dict(res_4df)
        df_result.to_csv(os.path.join(self.path_save_res,'ttest.csv'))


def get_stats(stats_type, data1, data2):
    '''
    Args: type of stats, data
    Return: float
    '''
    if stats_type == 'mean':
        return (stats.tmean(data1), stats.tmean(data2)), ('mean', 'mean')
    if stats_type == 'std':
        return (stats.tstd(data1), stats.tstd(data2)), ('std', 'std')
    elif stats_type == 'kurtosis':
        return (stats.kurtosis(data1), stats.kurtosis(data2)), ('kurtosis', 'kurtosis')
    elif stats_type == 'skewness':
        return (stats.skew(data1), stats.skew(data2)), ('skewness', 'skewness')
    elif stats_type == 'TTest':
        return stats.ttest_ind(data1, data2, equal_var = True), ('t','p')
    elif stats_type == 'Welch':
        return stats.ttest_ind(data1, data2, equal_var = False), ('t','p')
    elif stats_type == 'ANOVA':
        return stats.f_oneway(data1, data2), ('t','p') #One way ANOVA, checks the variance
    elif stats_type == 'Bartlett':
        return stats.bartlett(data1, data2), ('t','p') # Bartlett, tests the null hypothesis
    elif stats_type == 'MannWhitneyu':
        try:
            return stats.mannwhitneyu(data1, data2), ('u','p')
        except ValueError:
            return (0,0), ('h','p')
    elif stats_type == 'Kruskal':
        try:
            return stats.kruskal(data1, data2), ('h','p')
        except ValueError:
            return (0,0), ('h','p')


def mkstatisticsf(df_4stats,
                groups,
                group_col,
                path2save,
                make_with_colors = True):
    '''Creates discriptive statistical file for publication,
        based on provided pandas.DataFrame
        Works only on 2 groups
    Args: df_4stats: pandas.DataFrame
        group: list/ tuple of groups as str/int
        group_col: str() column name in df_4stats that has the group names from group
        path_2save: abspath to save the descrptive files
        make_with_colors: will create an additional .xlsx file with 
                        colored significant results,
                        provided xlwt is installed
    Return:
        json file with results
        .csv file with results
        .xlsx file with results with red colored significant ones
    '''

    tab = Table()
    ls_tests = ('mean', 'std',
                'kurtosis', 'skewness',
                'TTest', 'Welch', 'ANOVA', 
                'Bartlett', 'MannWhitneyu', 'Kruskal')

    groups_df = dict()
    for group in groups:
        groups_df[group] = tab.get_df_per_parameter(df_4stats, group_col, group)

    stats_dic = dict()
    vals2chk = df_4stats.columns.tolist()
    if group_col in vals2chk:
        vals2chk.remove(group_col)

    cols2color_sig = list()
    groups = list(groups_df.keys())
    group1 = groups_df[groups[0]]
    group2 = groups_df[groups[1]]
    for test in ls_tests:
        for val in vals2chk:
            results, params = get_stats(test, group1[val], group2[val])       
            if test in ('mean', 'std', 'kurtosis', 'skewness'):
                key1 = f'{groups[0]}, {params[0]}'
                key2 = f'{groups[1]}, {params[0]}'
            else:
                key1 = f'{test}, {params[0]}'
                key2 = f'{test}, {params[1]}'
                cols2color_sig.append(key2)
            for key in (key1, key2):
                if key not in stats_dic:
                    stats_dic[key] = dict()
            stats_dic[key1][val] = f'{results[0]}'
            stats_dic[key2][val] = f'{results[1]}'

    df = tab.create_df_from_dict(stats_dic)
    tab.save_df(df, os.path.join(path2save, 'stats_general.csv'), sheet_name = 'stats')
    utilities.save_json(stats_dic, os.path.join(path2save,'stats_general.json'))
    if make_with_colors:
        save_2xlsx_with_colors(df, path2save = path2save,
                           cols2color_sig = cols2color_sig)


def chk_if_module_ready(module):
    ls_import = list()
    try:
        ls_import.append(__import__(module))
        return True
    except ImportError:
        return False

def save_2xlsx_with_colors(df, path2save,
                           file_name = 'stats_general.xlsx',
                           cols2color_sig = [],
                           sig_thresh = 0.05,
                           nr_digits = 6):
    if not chk_if_module_ready('xlwt'):
        print('xlwt not installed, cannot continue saving to an xlsx file with colors')
        pass
    else:
        print('creating file xlsx with colors')
        import xlwt
        w=xlwt.Workbook()
        xlwt.add_palette_colour("custom_colour", 0x21)
        w.set_colour_RGB(0x21, 251, 228, 228)
        style = xlwt.easyxf('pattern: pattern solid, fore_colour custom_colour')
        style_bold = xlwt.easyxf('font: bold on; font: color blue')
        style_significant = xlwt.easyxf('font: bold on; font: color red')
        ws = w.add_sheet('stats_general', cell_overwrite_ok=True)
        icol = 0
        for col in df.columns.tolist():
            ws.write(0, icol+1, col, style_bold)
            if not cols2color_sig:
                if ', p' in col.lower() or col.lower().endswith('p'):
                    cols2color_sig.append(col)
            icol += 1
        irow = 0
        while irow < len(df.index):
            ws.write(irow+1, 0, df.index[irow], style_bold)
            for icol in range(0, len(df.columns.tolist())):
                col = df.columns.tolist()[icol]
                val = str(df.iloc[irow, icol])[:nr_digits]
                ws.write(irow+1, icol+1, val)
                if col in cols2color_sig and float(val) < sig_thresh:
                    ws.write(irow+1, icol+1, val, style_significant)                     
            irow += 1
        w.save(os.path.join(path2save, file_name))


def run_code_from_Emmanuelle(df_4stats, groups, group_col):
    df_new, stats_dic_4_json, cols2color_sig = mkstatisticsf_Emmanuelle(df_4stats, groups, group_col)

    path2save = os.getcwd()
    save_df_Emmanuelle(df_new, groups, stats_dic_4_json, cols2color_sig, path2save,
                      extensions = ['xlsx', 'csv', 'json'], make_with_colors = True)


def mkstatisticsf_Emmanuelle(df_4stats, groups, group_col,):
    '''Creates discriptive statistical file for publication,
        based on provided pandas.DataFrame
        Works only on 2 groups
        author: version adjusted by Emmanuelle Mazur-Lainé 202206
    Args: df_4stats: pandas.DataFrame
        group: list/ tuple of groups as str/int
        group_col: column name in df_4stats that has the group names from group
        path_2save: abspath to save the descriptive files
        make_with_colors: will create an additional .xlsx file with 
                        colored significant results,
                        provided xlwt is installed
    Return:
        json file with results
        .csv file with results
        .xlsx file with results with red colored significant ones
    '''
    df_4stats = df_4stats.astype(float)

    tab = Table()

    groups_df = dict()
    for group in groups:
        groups_df[group] = tab.get_df_per_parameter(df_4stats, group_col, group)

    stats_dic = dict()
    vals2chk = df_4stats.columns.tolist()

    if group_col in vals2chk:
        vals2chk.remove(group_col)

    cols2color_sig = list()
    groups = list(groups_df.keys())

    ################################
    if len(groups) == 1:
        ls_tests = ('mean', 'std', 'kurtosis', 'skewness')
    elif len(groups) <= 2 :
        ls_tests = ('mean', 'std',
                'kurtosis', 'skewness',
                'TTest', 'Welch', 'MannWhitneyu')
    elif len(groups) > 2 :
        ls_tests = ('mean', 'std',
                'kurtosis', 'skewness',
                'Bartlett', 'Kruskal', 'ANOVA')


    for test in ls_tests:
        for val in vals2chk:
            values_per_gr = []
            for i in range(0, len(groups)):
                gr_i = groups_df[groups[i]][val].values
                arr = np.array(gr_i)
                arr_without_nan = arr[np.logical_not(np.isnan(arr))]
                values_per_gr.append(arr_without_nan)

            results, params = get_stats_Emmanuelle(test, groups, values_per_gr)

            if test == 'mean' :
                for i in range(len(groups)):
                    for tst in ('mean', 'std'):
                        results, params = get_stats_Emmanuelle(tst, groups, values_per_gr)
                        key = f'{groups[i]}, {params}'
                        if key not in stats_dic:
                            stats_dic[key] = dict()
                        stats_dic[key][val] = f'{results[i]}'


            if test in ('kurtosis', 'skewness'):
                for i in range(len(groups)) :
                    key = f'{groups[i]}, {params}'
                    if key not in stats_dic :
                        stats_dic[key] = dict()
                    stats_dic[key][val] = f'{results[i]}'

            elif test in ('TTest', 'Welch',
                'Bartlett', 'MannWhitneyu', 'Kruskal', 'ANOVA'):
                for i in range(len(groups)) :
                    key1 = f'{test}, {params[0]}'
                    key2 = f'{test}, {params[1]}'
                    for key in (key1, key2) :
                        if key not in stats_dic :
                            stats_dic[key] = dict()
                    stats_dic[key1][val] = f'{results[0]}'
                    stats_dic[key2][val] = f'{results[1]}'

                cols2color_sig.append(key2)

    df = tab.create_df_from_dict(stats_dic)
    df = df.astype(float)

    # Creating new adjusted DataFrame with sub-indexes
    ls_tests_dup = []
    ls_param = []
    ls_keys = list(stats_dic.keys())

    mean_gr_done = False

    for test in ls_tests:
        if test in ('mean', 'std'):
            for i in range(0, len(groups)):
                ls_tests_dup.append('mean/std')
            if mean_gr_done == False:
                for i in range (len(groups)) :
                    ls_param.append('gr' + str(i + 1) + ' (val=' + f'{groups[i]}' + ')')
                    ls_param.append('gr' + str(i + 1) + ' (val=' + f'{groups[i]}' + ')')
                    mean_gr_done = True

        elif test in ('kurtosis', 'skewness'):
            for i in range(0, len(groups)) :
                ls_tests_dup.append(test)
                ls_param.append('gr' + str(i+1) + ' (val=' + f'{groups[i]}' + ')')

        elif test in ('TTest', 'Welch',
                'Bartlett', 'MannWhitneyu', 'Kruskal', 'ANOVA') :
            ls_tests_dup.append(test)
            ls_tests_dup.append(test)

    for key in ls_keys[4*(len(groups)):] :
        ls_param.append((str(key))[-1])


    col = [ls_tests_dup, ls_param]
    tuples = list(zip(*col))

    df_new = pd.DataFrame(df.values, index = pd.Index(df.index), columns = pd.MultiIndex.from_tuples(tuples))

    df_new = df_new.round(3)


    return df_new, stats_dic, cols2color_sig


def get_stats_Emmanuelle(stats_type, groups, *data):
    '''
        author: version adjusted by Emmanuelle Mazur-Lainé 202206
    Args: type of stats, data
    Return: float
    '''

    data = data[0]
    nbr_gr = len(groups)

    if stats_type == 'mean':
        res_stats = ()
        for group in data:
            res = stats.tmean(group)
            res_stats += (res,)
        return (res_stats), 'mean'
    if stats_type == 'std':
        res_stats = ()
        for group in data:
            res = stats.tstd(group)
            res_stats += (res,)
        return res_stats, 'std'
    elif stats_type == 'kurtosis':
        res_stats = ()
        for group in data:
            res = stats.kurtosis(group)
            res_stats += (res,)
        return res_stats, 'kurtosis'
    elif stats_type == 'skewness':
        res_stats = ()
        for group in data:
            res = stats.skew(group)
            res_stats += (res,)
        return res_stats, 'skewness'


    elif stats_type == 'TTest':
        return stats.ttest_ind(data[0], data[1], equal_var = True), ('t','p')
    elif stats_type == 'Welch':
        return stats.ttest_ind(data[0], data[1], equal_var = False), ('t','p')
    elif stats_type == 'MannWhitneyu':
        try:
            return stats.mannwhitneyu(data[0], data[1]), ('u','p')
        except ValueError:
            return (0,0), ('h','p')

    ########### RESTE À TROUVER COMMENT METTRE TOUS LES GROUPES
    # EN PRAMÈTRES DES TESTS BARTLETT, KRUSKAL ET ANOVA####

    elif stats_type == 'Bartlett':
        return stats.bartlett(*data), ('t','p') # Bartlett, tests the null hypothesis
    elif stats_type == 'Kruskal':
        try:
            return stats.kruskal(*data), ('h','p')
        except ValueError:
            return (0,0), ('h','p')
    elif stats_type == 'ANOVA':
        return stats.f_oneway(*data), ('t','p') #One way ANOVA, checks the variance


def f_coord(nbr_col):
    full_series = nbr_col // 26
    alpha = string.ascii_uppercase
    ls_coord = list(alpha)
    for i in range(full_series):
        for j in range(len(alpha)):
            ls_coord.append(str(alpha[i] + alpha[j]))

    return ls_coord

def save_2xlsx_with_colors_Emmanuelle(df, file_2get, path2save,
                           file_name,
                           cols2color_sig = [],
                           sig_thresh = 0.05):
    import openpyxl
    from openpyxl.styles import Font
    from openpyxl.cell import Cell

    file = openpyxl.load_workbook(file_2get)
    sheet = file.worksheets[0]
    sheet['O1'].font = Font(color = 'FFFF0000')

    nbr_col = len(df.columns)+1
    ls_coord = f_coord(nbr_col)
    i = 0
    icol = ls_coord[0]
    ls_sig = []
    for col in sheet.iter_cols(values_only=True):
        irow = 1
        if col[1] == 'p':
            for cell in col:
                cell_coord = ''.join((f'{icol}', f'{irow}'))
                value = sheet.cell(irow,i+1).value

                if type(value) is float and value <= sig_thresh:
                    ls_sig.append(value)
                    sheet[cell_coord].font = Font(color='FFFF0000')

                irow += 1

        i+=1
        icol = ls_coord[i]

    file.save(os.path.join(path2save, file_name))

########################################################################
# Fonction qui prend le DataFrame retourné par mkstatisticsf (ou n'importe quel DF) et qui crée fichiers .xslx, etc.
# Fonction sera appelée dans run

def save_df_Emmanuelle(df, groups, stats_dic, cols2color_sig, path2save, make_with_colors, extensions = ('xlsx', 'csv', 'json')) :

    if 'xlsx' in extensions :
        import openpyxl
        import string
        df.to_excel('stats_new.xlsx')
        ########## MERGE MEAN/STD SUB-INDEXES ################
        file = openpyxl.load_workbook('stats_new.xlsx')
        sheet = file['Sheet1']
        alpha = string.ascii_uppercase
        for ltr in range(len(alpha))[1:(2 * len(groups)) + 1:2]:
            cell1, cell2 = alpha[ltr] + str(2), alpha[ltr + 1] + str(2)
            sheet.merge_cells(str(cell1 + ':' + cell2))
        file.save('stats_new.xlsx')

    if 'json' in extensions :
        utilities.save_json(stats_dic, os.path.join(path2save, 'stats.json'))

    if 'csv' in extensions :
        tab = Table()
        tab.save_df(df, os.path.join(path2save, 'stats_new.csv'), sheet_name='stats')

    if make_with_colors:
        save_2xlsx_with_colors_Emmanuelle(df, 'stats_new.xlsx', path2save, 'stats_wcolors.xlsx',
                               cols2color_sig=cols2color_sig)




# def mkstatisticsf(df, group_col, GLM_dir):

#     print('Creating the group_statistics.xls file with statistical results for groups')
#     print(df.columns.tolist()[:5])
#     '''creating lists with the row limits for each group and the number of cotrasts'''
#     lsgroups = df[group_col]
#     groups = []
#     for x in lsgroups:
#             if x not in groups:
#                 groups.append(x)
#     lengroups = []
#     for i in range(0,len(groups)):
#         lengroups.append(0)
#     for x in lsgroups:
#         for group in groups:
#             if x == group:
#                 lengroups[groups.index(group)] += 1
#     # lengroups = [len(list(group)) for key, group in groupby(lsgroups)] #not working, don't know why
#     rownr = [0]
#     k = 0
#     v = 0
#     while k<len(lengroups):
#             v = lengroups[k]+v
#             rownr.append(v)
#             k+=1

#     NrOfGroupContrasts = 0
#     l = 0
#     while l<len(groups)-1:
#         t = l+1
#         while t<len(groups):
#             NrOfGroupContrasts += 1
#             t += 1
#         l += 1
#     #lists
#     lsstatsres = ('ttest t','ttest p','anova t','anova p','Bartlett t','Bartlett p',
#                   'MannWhitneyu u','MannWhitneyu p', 'Kruskal H','Kruskal p') #list of statistical results to be presented

#     '''creating the file with statistical results'''

#     w=xlwt.Workbook()
#     xlwt.add_palette_colour("custom_colour", 0x21)
#     w.set_colour_RGB(0x21, 251, 228, 228)
#     style = xlwt.easyxf('pattern: pattern solid, fore_colour custom_colour')
#     style_bold = xlwt.easyxf('font: bold on; font: color blue')
#     style_significant = xlwt.easyxf('font: bold on; font: color red')

#     ws = w.add_sheet('stats')
#     df.drop(['id',group_col], axis=1, inplace=True)
#     colnamesdata = df.columns.values.tolist()
    
#     base = 0

#     '''writing the first column with the names of the statistical tests'''
#     n = 0
#     row = base+2
#     col = base    
#     while n<NrOfGroupContrasts:
#             n1 = 0
#             while n1<len(lsstatsres):
#                 res = lsstatsres[n1]
#                 ws.write(row, col, res)
#                 n1 += 1
#                 row += 1
#             n += 1
#             row += 2
    
#     '''writing the results for each data column and each group contrast'''
#     colname = 0
#     while colname<len(colnamesdata):
#             row = base
#             col = colname + 1
#             ws.write(row, col, colnamesdata[colname])
        
#             row = base +1
#             l = 0
#             while l<len(groups)-1:
#                     t=l+1
#                     group2compare1 = df[colnamesdata[colname]][rownr[l]:rownr[t]]
#                                         #mk a list with the means and std of each groups

#                     '''writing the results for each data column and each group contrast'''
#                     while t<len(rownr)-1:
#                         grpcontrast=(groups[l]+'vs'+groups[t])
#                         group2compare2 = df[colnamesdata[colname]][rownr[t]:rownr[t+1]]
#                         tls = []
#                         pls = []
#                         ttestt, ttestp = stats.ttest_ind(group2compare1,group2compare2) #two-sided test 2 independent samples>
#                         anovat, anovap = stats.f_oneway(group2compare1,group2compare2) #One way Anova checks if the variance >
#                         bartt, bartp = stats.bartlett(group2compare1,group2compare2) #Bartlett’s test tests the null hypothes>
#                         try:
#                             manu, manp = stats.mannwhitneyu(group2compare1,group2compare2) #
#                         except ValueError:
#                             manu, manp = (0, 0)
#                         try:
#                             kruskh, kruskp = stats.kruskal(group2compare1,group2compare2) #
#                         except ValueError:
#                             kruskh, kruskp = (0, 0)
#                         tls = (ttestt, anovat, bartt, manu, kruskh)
#                         pls = (ttestp, anovap, bartp, manp, kruskp)
#                        #pearsoncorr, pearsonp = stats.pearsonr(group2compare1,group2compare2) #Calculates a Pearson correlati>
#                        #spearmancorr, spearmanp = stats.spearmanr(group2compare1,group2compare2) #Calculates a Spearman rank->
                    
#                         ws.write(row, col, grpcontrast, style_bold)
                    
#                         reslsnr = 0
#                         row += 1
#                         while reslsnr<len(tls):
#                             ws.write(row, col, tls[reslsnr])
#                             if pls[reslsnr]<0.05:
#                                 ws.write(row+1, col, pls[reslsnr], style_significant)
#                             else:
#                                 ws.write(row+1, col, pls[reslsnr])
#                             reslsnr += 1
#                             row += 2

#                         t=t+1
#                         row += 1
                    
#                     l=l+1

#             '''creating the plots for all groups for each column/value'''
#             #make a plot with means and std for all groups for each value
#             colname += 1

#     w.save(GLM_dir+'results/group_statistics.xls')
#     print('FINISHED creating General Statistics files')
